"""
Smart Attendance System — Streamlit Admin Dashboard
====================================================
Run with:  streamlit run dashboard.py

Pages:
  📊 Overview          — today's stats + KPI cards
  📅 Daily Report      — attendance for any chosen date
  👤 Student Detail    — per-student history & charts
  📋 Full Records      — searchable/filterable full table
  📥 Export            — download filtered Excel / CSV
  🗑️ Delete Attendance — remove attendance records by date
  ⚙️ Manage Students   — delete / rename enrolled students
"""

import os
import sys
import pickle
import io
from datetime import datetime, date, timedelta

import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from openpyxl import load_workbook

# ── Resolve attendance_system package path ──────────────────────────────────
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, BASE_DIR)

from config import (
    ATTENDANCE_FILE, LABEL_MAP_FILE, ATTENDANCE_DIR, KNOWN_FACES_DIR
)

# ═══════════════════════════════════════════════════════════════════════════════
#  PAGE CONFIG
# ═══════════════════════════════════════════════════════════════════════════════
st.set_page_config(
    page_title="Attendance Admin Dashboard",
    page_icon="📷",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ═══════════════════════════════════════════════════════════════════════════════
#  CUSTOM CSS — dark glassmorphism theme matching the Tkinter app palette
# ═══════════════════════════════════════════════════════════════════════════════
st.markdown("""
<style>
/* ── Root palette ─────────────────────────────────────────────── */
:root {
    --bg:       #1a1a2e;
    --surface:  #0f3460;
    --accent:   #16213e;
    --green:    #00C853;
    --orange:   #FF6D00;
    --red:      #FF1744;
    --text:     #e0e0e0;
    --muted:    #8888aa;
}

/* ── App background ────────────────────────────────────────────── */
.stApp {
    background: linear-gradient(135deg, #1a1a2e 0%, #0f3460 60%, #16213e 100%);
    color: var(--text);
}
.block-container { padding-top: 1.5rem; }

/* ── Sidebar ───────────────────────────────────────────────────── */
[data-testid="stSidebar"] {
    background: rgba(15, 52, 96, 0.85);
    backdrop-filter: blur(12px);
    border-right: 1px solid rgba(255,255,255,0.08);
}
[data-testid="stSidebar"] * { color: var(--text) !important; }

/* ── Radio nav ─────────────────────────────────────────────────── */
[data-testid="stSidebar"] .stRadio label {
    font-size: 0.95rem;
    padding: 6px 12px;
    border-radius: 8px;
    transition: background 0.2s;
    cursor: pointer;
}
[data-testid="stSidebar"] .stRadio label:hover {
    background: rgba(255,255,255,0.08);
}

/* ── KPI cards ──────────────────────────────────────────────────── */
.kpi-card {
    background: rgba(22,33,62,0.75);
    border: 1px solid rgba(255,255,255,0.10);
    border-radius: 16px;
    padding: 22px 20px 18px 20px;
    text-align: center;
    backdrop-filter: blur(8px);
    transition: transform 0.2s, box-shadow 0.2s;
}
.kpi-card:hover {
    transform: translateY(-3px);
    box-shadow: 0 8px 32px rgba(0,0,0,0.4);
}
.kpi-icon  { font-size: 2rem; margin-bottom: 4px; }
.kpi-value { font-size: 2.4rem; font-weight: 800; color: #ffffff; line-height: 1.1; }
.kpi-label { font-size: 0.80rem; color: var(--muted); margin-top: 4px; letter-spacing: 0.05em; text-transform: uppercase; }
.kpi-sub   { font-size: 0.78rem; color: var(--green); margin-top: 6px; font-weight: 600; }

/* ── Section headings ───────────────────────────────────────────── */
.section-title {
    font-size: 1.15rem;
    font-weight: 700;
    color: #ffffff;
    margin-bottom: 0.5rem;
    padding-bottom: 6px;
    border-bottom: 2px solid rgba(0,200,83,0.4);
}

/* ── Delete-attendance danger card ─────────────────────────────── */
.danger-card {
    background: rgba(255,23,68,0.08);
    border: 1px solid rgba(255,23,68,0.30);
    border-radius: 14px;
    padding: 18px 20px;
    margin-bottom: 1rem;
}

/* ── Badge pills ────────────────────────────────────────────────── */
.badge-present {
    background: rgba(0,200,83,0.18);
    color: var(--green);
    border: 1px solid rgba(0,200,83,0.4);
    border-radius: 20px;
    padding: 2px 12px;
    font-size: 0.78rem;
    font-weight: 700;
}
.badge-absent {
    background: rgba(255,23,68,0.15);
    color: var(--red);
    border: 1px solid rgba(255,23,68,0.35);
    border-radius: 20px;
    padding: 2px 12px;
    font-size: 0.78rem;
    font-weight: 700;
}

/* ── Dataframe ──────────────────────────────────────────────────── */
[data-testid="stDataFrame"] {
    border-radius: 12px;
    overflow: hidden;
    border: 1px solid rgba(255,255,255,0.08);
}

/* ── Progress bar ───────────────────────────────────────────────── */
.stProgress > div > div { background: var(--green) !important; }

/* ── Divider ─────────────────────────────────────────────────────── */
hr { border-color: rgba(255,255,255,0.08) !important; }

/* ── Buttons ─────────────────────────────────────────────────────── */
.stButton button {
    background: var(--surface);
    color: #fff;
    border: 1px solid rgba(255,255,255,0.15);
    border-radius: 8px;
    transition: background 0.2s;
}
.stButton button:hover { background: #1e3a5f; }

/* ── Plotly chart bg ─────────────────────────────────────────────── */
.js-plotly-plot .plotly { border-radius: 14px; }

/* ── Alerts ──────────────────────────────────────────────────────── */
.stAlert { border-radius: 10px; }
</style>
""", unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════════════════════
#  DATA HELPERS
# ═══════════════════════════════════════════════════════════════════════════════


def _delete_attendance_records(target_date: date, student_names: list[str]) -> tuple[bool, str, int]:
    """
    Delete attendance rows from the Excel file for the given date and
    student name(s).  If student_names is empty, all rows for that date
    are removed (wipe entire day).

    Returns (success, message, rows_deleted).
    """
    if not os.path.exists(ATTENDANCE_FILE):
        return False, "Attendance file not found.", 0

    target_str = str(target_date)
    try:
        wb = load_workbook(ATTENDANCE_FILE)
        ws = wb.active

        rows_to_delete = []
        for row in ws.iter_rows(min_row=2):
            row_date = str(row[3].value).strip() if row[3].value else ""
            row_name = str(row[2].value).strip() if row[2].value else ""

            date_matches = (row_date == target_str)
            name_matches = (
                not student_names  # empty = delete whole day
                or any(row_name.lower() == n.lower() for n in student_names)
            )
            if date_matches and name_matches:
                rows_to_delete.append(row[0].row)

        if not rows_to_delete:
            return False, f"No matching records found for {target_date}.", 0

        # Delete in reverse order to preserve row indices
        for r in sorted(rows_to_delete, reverse=True):
            ws.delete_rows(r)

        # Re-number S.No column
        for idx, row in enumerate(ws.iter_rows(min_row=2), start=1):
            row[0].value = idx

        wb.save(ATTENDANCE_FILE)
        wb.close()
        return True, f"Deleted {len(rows_to_delete)} record(s) for {target_date}.", len(rows_to_delete)
    except Exception as e:
        return False, f"Failed to delete records: {e}", 0


def _rename_student(old_name: str, old_roll: str, new_name: str, new_roll: str) -> tuple[bool, str]:
    """
    Rename a student in:
      1. The label-map pickle (face_encoder data)
      2. The known_faces folder name
      3. The attendance Excel workbook (all historical rows)
    Returns (success, message).
    """
    import re, shutil
    from openpyxl import load_workbook as _lw

    def _san(v): return re.sub(r'[^\w\-. ]', '_', v).strip()

    # ── 1. Update label map ───────────────────────────────────────────────
    if not os.path.exists(LABEL_MAP_FILE):
        return False, "Label map not found."
    try:
        with open(LABEL_MAP_FILE, "rb") as f:
            label_map = pickle.load(f)
    except Exception as e:
        return False, f"Could not read label map: {e}"

    if old_name not in label_map:
        return False, f"Student '{old_name}' not found in label map."

    entry = label_map.pop(old_name)
    entry["roll_no"] = new_roll
    label_map[new_name] = entry

    with open(LABEL_MAP_FILE, "wb") as f:
        pickle.dump(label_map, f)

    # ── 2. Rename known_faces folder ──────────────────────────────────────
    old_folder = os.path.join(KNOWN_FACES_DIR, f"{_san(old_roll)}_{_san(old_name)}")
    new_folder = os.path.join(KNOWN_FACES_DIR, f"{_san(new_roll)}_{_san(new_name)}")
    if os.path.exists(old_folder) and old_folder != new_folder:
        try:
            shutil.move(old_folder, new_folder)
        except Exception as e:
            return False, f"Folder rename failed: {e}"

    # ── 3. Patch attendance Excel ─────────────────────────────────────────
    if os.path.exists(ATTENDANCE_FILE):
        try:
            wb = _lw(ATTENDANCE_FILE)
            ws = wb.active
            for row in ws.iter_rows(min_row=2):
                if row[2].value and str(row[2].value).strip().lower() == old_name.lower():
                    row[2].value = new_name
                if row[1].value and str(row[1].value).strip().lower() == old_roll.lower():
                    row[1].value = new_roll
            wb.save(ATTENDANCE_FILE)
            wb.close()
        except Exception as e:
            return False, f"Attendance file update failed: {e}"

    return True, f"'{old_name}' → '{new_name}' updated successfully."


def _delete_student_dashboard(name: str, purge_attendance: bool = True) -> tuple[bool, str]:
    """
    Delete a student's face data and (optionally) purge their attendance rows
    from the Excel file so old records don't inflate present/absent counts.
    """
    try:
        from face_encoder import delete_student
        ok = delete_student(name)
        if not ok:
            return False, f"'{name}' was not found in the system."
    except Exception as e:
        return False, f"Delete failed: {e}"

    # ── Purge attendance rows for this student ──────────────────────────────
    if purge_attendance and os.path.exists(ATTENDANCE_FILE):
        try:
            from openpyxl import load_workbook as _lw
            wb = _lw(ATTENDANCE_FILE)
            ws = wb.active
            rows_to_delete = [
                row[0].row for row in ws.iter_rows(min_row=2)
                if row[2].value and str(row[2].value).strip().lower() == name.lower()
            ]
            for r in sorted(rows_to_delete, reverse=True):
                ws.delete_rows(r)
            for idx, row in enumerate(ws.iter_rows(min_row=2), start=1):
                row[0].value = idx
            wb.save(ATTENDANCE_FILE)
            wb.close()
            purge_note = f" Removed {len(rows_to_delete)} attendance record(s)."
        except Exception as e:
            purge_note = f" (Attendance purge failed: {e})"
    else:
        purge_note = ""

    return True, f"'{name}' deleted and model retrained.{purge_note}"


@st.cache_data(ttl=30)          # refresh every 30 seconds
def load_attendance_df() -> pd.DataFrame:
    """
    Load the master attendance Excel into a tidy DataFrame.
    This is the single source of truth — the dashboard always reads
    straight from the same file that main.py writes to.
    """
    if not os.path.exists(ATTENDANCE_FILE):
        return pd.DataFrame(columns=["S.No", "Roll No", "Student Name", "Date", "Time", "Status"])
    try:
        wb = load_workbook(ATTENDANCE_FILE, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        wb.close()
        df = pd.DataFrame(rows, columns=["S.No", "Roll No", "Student Name", "Date", "Time", "Status"])
        df = df.dropna(subset=["Student Name"])
        df["Date"] = pd.to_datetime(df["Date"], errors="coerce").dt.date
        df["Time"] = df["Time"].astype(str)
        return df
    except Exception as e:
        st.error(f"Could not load attendance file: {e}")
        return pd.DataFrame(columns=["S.No", "Roll No", "Student Name", "Date", "Time", "Status"])


@st.cache_data(ttl=30)
def load_enrolled_students() -> list[dict]:
    """Return list of {name, roll_no} from the label map pickle."""
    if not os.path.exists(LABEL_MAP_FILE):
        return []
    try:
        with open(LABEL_MAP_FILE, "rb") as f:
            label_map = pickle.load(f)
        return [{"name": n, "roll_no": d["roll_no"]} for n, d in label_map.items()]
    except Exception:
        return []


def get_student_stats(df: pd.DataFrame, enrolled: list[dict]):
    """
    Compute per-student attendance percentage across all session dates.
    Returns a DataFrame with columns: name, roll_no, present_days, total_days, pct.
    """
    all_dates = df["Date"].dropna().unique()
    total_days = len(all_dates)

    rows = []
    for s in enrolled:
        name = s["name"]
        roll = s["roll_no"]
        present = df[df["Student Name"].str.lower() == name.lower()]["Date"].nunique()
        pct = round(present / total_days * 100, 1) if total_days > 0 else 0.0
        rows.append({"Student Name": name, "Roll No": roll,
                     "Present": present, "Absent": total_days - present,
                     "Total Days": total_days, "Attendance %": pct})
    return pd.DataFrame(rows)


def pct_color(pct: float) -> str:
    if pct >= 75:
        return "#00C853"
    elif pct >= 50:
        return "#FF6D00"
    return "#FF1744"


# ═══════════════════════════════════════════════════════════════════════════════
#  SIDEBAR
# ═══════════════════════════════════════════════════════════════════════════════

with st.sidebar:
    st.markdown("## 📷 Attendance Admin")
    st.markdown("<small style='color:#8888aa'>Smart Attendance System</small>", unsafe_allow_html=True)
    st.divider()

    page = st.radio(
        "Navigate",
        [
            "📊 Overview",
            "📅 Daily Report",
            "👤 Student Detail",
            "📋 Full Records",
            "📥 Export",
            "🗑️ Delete Attendance",
            "⚙️ Manage Students",
        ],
        label_visibility="collapsed",
    )

    st.divider()

    if st.button("🔄 Refresh Data"):
        st.cache_data.clear()
        st.rerun()

    st.markdown(
        f"<small style='color:#8888aa'>Last refreshed:<br>{datetime.now().strftime('%H:%M:%S')}</small>",
        unsafe_allow_html=True,
    )
    st.divider()
    st.markdown(
        f"<small style='color:#8888aa'>📁 Data file:<br><code style='font-size:0.7rem'>"
        f"{os.path.basename(ATTENDANCE_FILE)}</code></small>",
        unsafe_allow_html=True,
    )


# ═══════════════════════════════════════════════════════════════════════════════
#  LOAD DATA (shared across all pages)
# ═══════════════════════════════════════════════════════════════════════════════

df_all     = load_attendance_df()
enrolled   = load_enrolled_students()
today_str  = date.today()
df_today   = df_all[df_all["Date"] == today_str]

total_enrolled = len(enrolled)
present_today  = df_today["Student Name"].nunique()
absent_today   = max(0, total_enrolled - present_today)
all_dates      = sorted(df_all["Date"].dropna().unique())
total_sessions = len(all_dates)


# ═══════════════════════════════════════════════════════════════════════════════
#  PAGE: OVERVIEW
# ═══════════════════════════════════════════════════════════════════════════════

if page == "📊 Overview":
    st.markdown("<h1 style='color:#fff;margin-bottom:0'>📊 Dashboard Overview</h1>", unsafe_allow_html=True)
    st.markdown(f"<p style='color:#8888aa;margin-top:4px'>{datetime.now().strftime('%A, %B %d %Y')}</p>", unsafe_allow_html=True)

    # ── KPI Cards ──────────────────────────────────────────────────────────
    c1, c2, c3, c4 = st.columns(4)
    pct_today = round(present_today / total_enrolled * 100) if total_enrolled else 0

    with c1:
        st.markdown(f"""
        <div class="kpi-card">
          <div class="kpi-icon">🎓</div>
          <div class="kpi-value">{total_enrolled}</div>
          <div class="kpi-label">Enrolled Students</div>
        </div>""", unsafe_allow_html=True)

    with c2:
        st.markdown(f"""
        <div class="kpi-card">
          <div class="kpi-icon">✅</div>
          <div class="kpi-value" style="color:#00C853">{present_today}</div>
          <div class="kpi-label">Present Today</div>
          <div class="kpi-sub">{pct_today}% attendance rate</div>
        </div>""", unsafe_allow_html=True)

    with c3:
        st.markdown(f"""
        <div class="kpi-card">
          <div class="kpi-icon">❌</div>
          <div class="kpi-value" style="color:#FF1744">{absent_today}</div>
          <div class="kpi-label">Absent Today</div>
        </div>""", unsafe_allow_html=True)

    with c4:
        st.markdown(f"""
        <div class="kpi-card">
          <div class="kpi-icon">📅</div>
          <div class="kpi-value">{total_sessions}</div>
          <div class="kpi-label">Total Sessions</div>
        </div>""", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # ── Charts row ──────────────────────────────────────────────────────────
    col_left, col_right = st.columns([1.6, 1], gap="large")

    with col_left:
        st.markdown("<div class='section-title'>📈 Daily Attendance Trend</div>", unsafe_allow_html=True)
        if len(all_dates) > 0:
            trend_data = (
                df_all.groupby("Date")["Student Name"]
                .nunique()
                .reset_index()
                .rename(columns={"Student Name": "Present"})
            )
            trend_data["Date"] = pd.to_datetime(trend_data["Date"])
            fig = px.area(
                trend_data, x="Date", y="Present",
                color_discrete_sequence=["#00C853"],
                template="plotly_dark",
            )
            fig.update_traces(
                fill="tozeroy",
                fillcolor="rgba(0,200,83,0.12)",
                line=dict(color="#00C853", width=2.5),
                mode="lines+markers",
                marker=dict(size=6, color="#00C853"),
            )
            fig.update_layout(
                paper_bgcolor="rgba(15,52,96,0.5)",
                plot_bgcolor="rgba(0,0,0,0)",
                font=dict(color="#e0e0e0"),
                margin=dict(l=10, r=10, t=10, b=10),
                xaxis=dict(gridcolor="rgba(255,255,255,0.05)"),
                yaxis=dict(gridcolor="rgba(255,255,255,0.05)", title="Students Present"),
                showlegend=False,
                height=280,
            )
            st.plotly_chart(fig, use_container_width=True)
        else:
            st.info("No attendance records yet.")

    with col_right:
        st.markdown("<div class='section-title'>🍩 Today's Snapshot</div>", unsafe_allow_html=True)
        if total_enrolled > 0:
            fig2 = go.Figure(go.Pie(
                labels=["Present", "Absent"],
                values=[present_today, absent_today],
                hole=0.62,
                marker=dict(colors=["#00C853", "#FF1744"],
                            line=dict(color="rgba(0,0,0,0)", width=0)),
                textinfo="label+percent",
                textfont=dict(color="#e0e0e0", size=13),
            ))
            fig2.update_layout(
                paper_bgcolor="rgba(15,52,96,0.5)",
                font=dict(color="#e0e0e0"),
                showlegend=False,
                margin=dict(l=10, r=10, t=10, b=10),
                height=280,
                annotations=[dict(
                    text=f"<b>{pct_today}%</b>",
                    x=0.5, y=0.5,
                    font=dict(size=26, color="#ffffff"),
                    showarrow=False,
                )],
            )
            st.plotly_chart(fig2, use_container_width=True)
        else:
            st.info("No students enrolled yet.")

    # ── Today's present list ─────────────────────────────────────────────────
    st.markdown("<div class='section-title'>🕐 Today's Check-ins</div>", unsafe_allow_html=True)
    if not df_today.empty:
        display = df_today[["Roll No", "Student Name", "Time", "Status"]].reset_index(drop=True)
        st.dataframe(
            display.style.map(
                lambda v: "color: #00C853; font-weight:600" if v == "Present" else "",
                subset=["Status"],
            ),
            use_container_width=True,
            hide_index=True,
        )
    else:
        st.info("No attendance marked yet today.")

    # ── Attendance % per student bar chart ──────────────────────────────────
    if enrolled and total_sessions > 0:
        st.markdown("<br><div class='section-title'>📊 Attendance Rate by Student</div>", unsafe_allow_html=True)
        stats_df = get_student_stats(df_all, enrolled)
        stats_df = stats_df.sort_values("Attendance %", ascending=True)

        colors = [pct_color(p) for p in stats_df["Attendance %"]]
        fig3 = go.Figure(go.Bar(
            x=stats_df["Attendance %"],
            y=stats_df["Student Name"],
            orientation="h",
            marker=dict(color=colors, line=dict(width=0)),
            text=[f"{p}%" for p in stats_df["Attendance %"]],
            textposition="outside",
            textfont=dict(color="#e0e0e0", size=12),
        ))
        fig3.add_vline(x=75, line_dash="dash", line_color="#FF6D00",
                       annotation_text="75% threshold", annotation_font_color="#FF6D00")
        fig3.update_layout(
            paper_bgcolor="rgba(15,52,96,0.5)",
            plot_bgcolor="rgba(0,0,0,0)",
            font=dict(color="#e0e0e0"),
            xaxis=dict(range=[0, 115], gridcolor="rgba(255,255,255,0.05)", title="Attendance %"),
            yaxis=dict(gridcolor="rgba(255,255,255,0)", title=""),
            margin=dict(l=10, r=40, t=10, b=10),
            height=max(280, len(enrolled) * 42),
        )
        st.plotly_chart(fig3, use_container_width=True)


# ═══════════════════════════════════════════════════════════════════════════════
#  PAGE: DAILY REPORT
# ═══════════════════════════════════════════════════════════════════════════════

elif page == "📅 Daily Report":
    st.markdown("<h1 style='color:#fff;margin-bottom:0'>📅 Daily Report</h1>", unsafe_allow_html=True)
    st.markdown("<p style='color:#8888aa'>Select a date to view that session's attendance.</p>", unsafe_allow_html=True)

    # Date picker
    min_date = min(all_dates) if all_dates else date.today()
    selected_date = st.date_input(
        "Select date",
        value=date.today(),
        min_value=min_date,
        max_value=date.today(),
    )

    df_day = df_all[df_all["Date"] == selected_date]
    present_names = set(df_day["Student Name"].str.lower())
    enrolled_names = [(s["name"], s["roll_no"]) for s in enrolled]

    col1, col2, col3 = st.columns(3)
    day_present = df_day["Student Name"].nunique()
    day_absent  = max(0, total_enrolled - day_present)
    day_pct     = round(day_present / total_enrolled * 100) if total_enrolled else 0

    with col1:
        st.markdown(f"""
        <div class="kpi-card">
          <div class="kpi-value" style="color:#00C853">{day_present}</div>
          <div class="kpi-label">✅ Present</div>
        </div>""", unsafe_allow_html=True)
    with col2:
        st.markdown(f"""
        <div class="kpi-card">
          <div class="kpi-value" style="color:#FF1744">{day_absent}</div>
          <div class="kpi-label">❌ Absent</div>
        </div>""", unsafe_allow_html=True)
    with col3:
        st.markdown(f"""
        <div class="kpi-card">
          <div class="kpi-value">{day_pct}%</div>
          <div class="kpi-label">📈 Attendance Rate</div>
        </div>""", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # Full class list with present/absent status
    st.markdown("<div class='section-title'>📋 Full Class List</div>", unsafe_allow_html=True)

    if enrolled_names:
        table_rows = []
        for name, roll in enrolled_names:
            is_present = name.lower() in present_names
            row_time = ""
            if is_present:
                match = df_day[df_day["Student Name"].str.lower() == name.lower()]
                row_time = match["Time"].values[0] if not match.empty else ""
            table_rows.append({
                "Roll No":      roll,
                "Student Name": name,
                "Status":       "✅ Present" if is_present else "❌ Absent",
                "Time":         row_time,
            })

        report_df = pd.DataFrame(table_rows)
        st.dataframe(
            report_df.style.map(
                lambda v: "color:#00C853;font-weight:600" if "Present" in str(v)
                          else ("color:#FF1744;font-weight:600" if "Absent" in str(v) else ""),
                subset=["Status"],
            ),
            use_container_width=True,
            hide_index=True,
        )

        # Download daily report
        csv_bytes = report_df.to_csv(index=False).encode()
        st.download_button(
            "📥 Download This Day's Report (CSV)",
            data=csv_bytes,
            file_name=f"attendance_{selected_date}.csv",
            mime="text/csv",
        )
    else:
        if df_day.empty:
            st.info(f"No attendance records for {selected_date}.")
        else:
            st.dataframe(df_day[["Roll No", "Student Name", "Time", "Status"]], use_container_width=True, hide_index=True)


# ═══════════════════════════════════════════════════════════════════════════════
#  PAGE: STUDENT DETAIL
# ═══════════════════════════════════════════════════════════════════════════════

elif page == "👤 Student Detail":
    st.markdown("<h1 style='color:#fff;margin-bottom:0'>👤 Student Detail</h1>", unsafe_allow_html=True)
    st.markdown("<p style='color:#8888aa'>Individual attendance history and statistics.</p>", unsafe_allow_html=True)

    if not enrolled:
        st.warning("No students enrolled yet.")
    else:
        student_names = [s["name"] for s in enrolled]
        selected_name = st.selectbox("Select student", student_names)
        student = next((s for s in enrolled if s["name"] == selected_name), None)

        if student:
            st.divider()
            # Stats
            s_df = df_all[df_all["Student Name"].str.lower() == selected_name.lower()]
            dates_present = sorted(s_df["Date"].dropna().unique())
            all_session_dates = sorted(df_all["Date"].dropna().unique())
            total = len(all_session_dates)
            present = len(dates_present)
            absent = total - present
            pct = round(present / total * 100, 1) if total > 0 else 0.0
            color = pct_color(pct)

            c1, c2, c3, c4 = st.columns(4)
            with c1:
                st.markdown(f"""<div class="kpi-card">
                  <div class="kpi-value">{student['roll_no']}</div>
                  <div class="kpi-label">🪪 Roll No</div>
                </div>""", unsafe_allow_html=True)
            with c2:
                st.markdown(f"""<div class="kpi-card">
                  <div class="kpi-value" style="color:#00C853">{present}</div>
                  <div class="kpi-label">✅ Days Present</div>
                </div>""", unsafe_allow_html=True)
            with c3:
                st.markdown(f"""<div class="kpi-card">
                  <div class="kpi-value" style="color:#FF1744">{absent}</div>
                  <div class="kpi-label">❌ Days Absent</div>
                </div>""", unsafe_allow_html=True)
            with c4:
                st.markdown(f"""<div class="kpi-card">
                  <div class="kpi-value" style="color:{color}">{pct}%</div>
                  <div class="kpi-label">📈 Attendance Rate</div>
                </div>""", unsafe_allow_html=True)

            st.markdown("<br>", unsafe_allow_html=True)

            # Attendance % gauge
            col_gauge, col_cal = st.columns([1, 1.6], gap="large")

            with col_gauge:
                st.markdown("<div class='section-title'>🎯 Attendance Gauge</div>", unsafe_allow_html=True)
                fig_g = go.Figure(go.Indicator(
                    mode="gauge+number+delta",
                    value=pct,
                    delta={"reference": 75, "increasing": {"color": "#00C853"}, "decreasing": {"color": "#FF1744"}},
                    gauge={
                        "axis": {"range": [0, 100], "tickcolor": "#8888aa"},
                        "bar":  {"color": color, "thickness": 0.28},
                        "bgcolor": "rgba(0,0,0,0)",
                        "bordercolor": "rgba(255,255,255,0.1)",
                        "steps": [
                            {"range": [0, 50],   "color": "rgba(255,23,68,0.15)"},
                            {"range": [50, 75],  "color": "rgba(255,109,0,0.15)"},
                            {"range": [75, 100], "color": "rgba(0,200,83,0.15)"},
                        ],
                        "threshold": {"line": {"color": "#FF6D00", "width": 3}, "thickness": 0.85, "value": 75},
                    },
                    number={"suffix": "%", "font": {"color": color, "size": 42}},
                ))
                fig_g.update_layout(
                    paper_bgcolor="rgba(15,52,96,0.5)",
                    font=dict(color="#e0e0e0"),
                    height=280,
                    margin=dict(l=20, r=20, t=30, b=10),
                )
                st.plotly_chart(fig_g, use_container_width=True)

                if pct >= 75:
                    st.success("✅ Attendance is satisfactory (≥ 75%)")
                elif pct >= 50:
                    st.warning("⚠️ Attendance is below 75% threshold!")
                else:
                    st.error("🚨 Critical — attendance below 50%!")

            with col_cal:
                st.markdown("<div class='section-title'>📅 Session History</div>", unsafe_allow_html=True)
                if all_session_dates:
                    cal_df = pd.DataFrame({
                        "Date":   [str(d) for d in all_session_dates],
                        "Status": ["Present" if d in dates_present else "Absent"
                                   for d in all_session_dates],
                    })
                    cal_df["Date_dt"] = pd.to_datetime(cal_df["Date"])
                    fig_cal = px.bar(
                        cal_df,
                        x="Date",
                        y=[1] * len(cal_df),
                        color="Status",
                        color_discrete_map={"Present": "#00C853", "Absent": "#FF1744"},
                        template="plotly_dark",
                    )
                    fig_cal.update_layout(
                        paper_bgcolor="rgba(15,52,96,0.5)",
                        plot_bgcolor="rgba(0,0,0,0)",
                        showlegend=True,
                        font=dict(color="#e0e0e0"),
                        yaxis=dict(visible=False),
                        xaxis=dict(title="", gridcolor="rgba(255,255,255,0.04)"),
                        bargap=0.15,
                        height=280,
                        margin=dict(l=10, r=10, t=10, b=10),
                        legend=dict(orientation="h", yanchor="bottom", y=1.02, x=0),
                    )
                    st.plotly_chart(fig_cal, use_container_width=True)

            # Detailed table
            st.markdown("<div class='section-title'>🕐 Check-in Log</div>", unsafe_allow_html=True)
            if not s_df.empty:
                st.dataframe(
                    s_df[["Date", "Time", "Status"]].sort_values("Date", ascending=False).reset_index(drop=True),
                    use_container_width=True,
                    hide_index=True,
                )
            else:
                st.info("No attendance records for this student.")


# ═══════════════════════════════════════════════════════════════════════════════
#  PAGE: FULL RECORDS
# ═══════════════════════════════════════════════════════════════════════════════

elif page == "📋 Full Records":
    st.markdown("<h1 style='color:#fff;margin-bottom:0'>📋 Full Attendance Records</h1>", unsafe_allow_html=True)
    st.markdown("<p style='color:#8888aa'>Browse, search, and filter all attendance data.</p>", unsafe_allow_html=True)

    if df_all.empty:
        st.info("No attendance records found.")
    else:
        # Filters
        col_f1, col_f2, col_f3 = st.columns(3)
        with col_f1:
            search_name = st.text_input("🔍 Search by name", placeholder="Type to filter…")
        with col_f2:
            unique_dates = sorted(df_all["Date"].dropna().unique(), reverse=True)
            date_options = ["All Dates"] + [str(d) for d in unique_dates]
            date_filter = st.selectbox("📅 Filter by date", date_options)
        with col_f3:
            enrolled_names_all = ["All Students"] + [s["name"] for s in enrolled]
            name_filter = st.selectbox("👤 Filter by student", enrolled_names_all)

        filtered = df_all.copy()
        if search_name:
            filtered = filtered[filtered["Student Name"].str.contains(search_name, case=False, na=False)]
        if date_filter != "All Dates":
            filtered = filtered[filtered["Date"].astype(str) == date_filter]
        if name_filter != "All Students":
            filtered = filtered[filtered["Student Name"].str.lower() == name_filter.lower()]

        st.markdown(f"<small style='color:#8888aa'>Showing {len(filtered)} records</small>", unsafe_allow_html=True)
        st.dataframe(
            filtered[["S.No", "Roll No", "Student Name", "Date", "Time", "Status"]]
            .sort_values("Date", ascending=False)
            .reset_index(drop=True),
            use_container_width=True,
            hide_index=True,
            height=480,
        )

        # Summary stats table
        if not filtered.empty and enrolled:
            st.markdown("<br><div class='section-title'>📊 Summary by Student (filtered)</div>", unsafe_allow_html=True)
            stats_df = get_student_stats(df_all, enrolled)
            if name_filter != "All Students":
                stats_df = stats_df[stats_df["Student Name"].str.lower() == name_filter.lower()]
            st.dataframe(stats_df.sort_values("Attendance %", ascending=False).reset_index(drop=True),
                         use_container_width=True, hide_index=True)


# ═══════════════════════════════════════════════════════════════════════════════
#  PAGE: EXPORT
# ═══════════════════════════════════════════════════════════════════════════════

elif page == "📥 Export":
    st.markdown("<h1 style='color:#fff;margin-bottom:0'>📥 Export Data</h1>", unsafe_allow_html=True)
    st.markdown("<p style='color:#8888aa'>Download attendance records in your preferred format.</p>", unsafe_allow_html=True)

    if df_all.empty:
        st.info("No data to export yet.")
    else:
        # ── Date range filter ────────────────────────────────────────────────
        st.markdown("<div class='section-title'>🗓️ Select Date Range</div>", unsafe_allow_html=True)
        col_d1, col_d2 = st.columns(2)
        min_d = min(all_dates) if all_dates else date.today()
        with col_d1:
            start_date = st.date_input("From", value=min_d, min_value=min_d, max_value=date.today())
        with col_d2:
            end_date = st.date_input("To", value=date.today(), min_value=min_d, max_value=date.today())

        export_df = df_all[(df_all["Date"] >= start_date) & (df_all["Date"] <= end_date)]

        st.markdown(f"**{len(export_df)} records** in selected range ({start_date} → {end_date})")

        col_e1, col_e2, col_e3 = st.columns(3)

        # ── CSV export ───────────────────────────────────────────────────────
        with col_e1:
            st.markdown("#### 📄 CSV")
            st.markdown("Lightweight, works with Excel / Google Sheets.")
            csv_data = export_df.to_csv(index=False).encode()
            st.download_button(
                "⬇️ Download CSV",
                data=csv_data,
                file_name=f"attendance_{start_date}_to_{end_date}.csv",
                mime="text/csv",
                use_container_width=True,
            )

        # ── Excel export ─────────────────────────────────────────────────────
        with col_e2:
            st.markdown("#### 📊 Excel (.xlsx)")
            st.markdown("Formatted workbook with styled headers.")
            buf = io.BytesIO()
            with pd.ExcelWriter(buf, engine="openpyxl") as writer:
                export_df.to_excel(writer, index=False, sheet_name="Attendance")
                if enrolled:
                    stats_df = get_student_stats(df_all, enrolled)
                    stats_df.to_excel(writer, index=False, sheet_name="Summary")
            buf.seek(0)
            st.download_button(
                "⬇️ Download Excel",
                data=buf,
                file_name=f"attendance_{start_date}_to_{end_date}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

        # ── Summary report ────────────────────────────────────────────────────
        with col_e3:
            st.markdown("#### 📋 Summary Report (CSV)")
            st.markdown("Attendance % per student across all sessions.")
            if enrolled:
                stats_df = get_student_stats(df_all, enrolled)
                summary_csv = stats_df.to_csv(index=False).encode()
                st.download_button(
                    "⬇️ Download Summary",
                    data=summary_csv,
                    file_name=f"attendance_summary_{date.today()}.csv",
                    mime="text/csv",
                    use_container_width=True,
                )
            else:
                st.info("No enrolled students.")

        # ── Preview ────────────────────────────────────────────────────────────
        st.divider()
        st.markdown("<div class='section-title'>👁️ Export Preview</div>", unsafe_allow_html=True)
        st.dataframe(
            export_df.head(20).reset_index(drop=True),
            use_container_width=True,
            hide_index=True,
        )
        if len(export_df) > 20:
            st.caption(f"Showing first 20 of {len(export_df)} rows.")


# ═══════════════════════════════════════════════════════════════════════════════
#  PAGE: DELETE ATTENDANCE
# ═══════════════════════════════════════════════════════════════════════════════

elif page == "🗑️ Delete Attendance":
    st.markdown("<h1 style='color:#fff;margin-bottom:0'>🗑️ Delete Attendance Records</h1>", unsafe_allow_html=True)
    st.markdown(
        "<p style='color:#8888aa'>Remove incorrect or erroneous attendance entries "
        "for any date. Changes are written directly to the Excel file.</p>",
        unsafe_allow_html=True,
    )

    if df_all.empty:
        st.info("No attendance records exist yet.")
    else:
        # ── Step 1: Pick a date ───────────────────────────────────────────────
        st.markdown("<div class='section-title'>📅 Step 1 — Choose a Date</div>", unsafe_allow_html=True)

        # Build selectable dates: today + all past dates that have records
        record_dates = sorted(df_all["Date"].dropna().unique(), reverse=True)
        date_options_del = [str(d) for d in record_dates]

        if not date_options_del:
            st.info("No dated records found in the file.")
            st.stop()

        selected_del_date_str = st.selectbox(
            "Select date to manage",
            date_options_del,
            help="Dates that have at least one attendance record are listed here.",
        )
        selected_del_date = date.fromisoformat(selected_del_date_str)

        # Show records for that date
        df_sel_day = df_all[df_all["Date"] == selected_del_date].copy()

        st.markdown(f"<br>", unsafe_allow_html=True)
        st.markdown(
            f"<div class='section-title'>📋 Records on {selected_del_date_str} "
            f"<span style='color:#8888aa;font-size:0.85rem'>({len(df_sel_day)} record(s))</span></div>",
            unsafe_allow_html=True,
        )

        if df_sel_day.empty:
            st.info(f"No records found for {selected_del_date_str}.")
            st.stop()

        # Display the day's records (read from Excel — always in sync)
        display_cols = ["Roll No", "Student Name", "Time", "Status"]
        st.dataframe(
            df_sel_day[display_cols].reset_index(drop=True),
            use_container_width=True,
            hide_index=True,
        )

        st.divider()

        # ── Step 2: Select scope ───────────────────────────────────────────────
        st.markdown("<div class='section-title'>🎯 Step 2 — Select Scope</div>", unsafe_allow_html=True)

        delete_mode = st.radio(
            "What do you want to delete?",
            [
                "🗑️ Delete attendance for specific student(s) on this date",
                "💣 Delete ALL attendance records for this entire date",
            ],
            key="del_att_mode",
        )

        students_to_delete: list[str] = []

        if "specific student" in delete_mode:
            present_on_day = sorted(df_sel_day["Student Name"].unique().tolist())
            students_to_delete = st.multiselect(
                "Select student(s) whose attendance to remove",
                present_on_day,
                help="Only students marked present on the selected date are shown.",
            )
            if not students_to_delete:
                st.info("Select at least one student to continue.")

        else:
            st.markdown(
                "<div class='danger-card'>⚠️ <b>Warning:</b> This will remove <b>all</b> "
                f"attendance entries for <b>{selected_del_date_str}</b> from the Excel file. "
                "This cannot be undone.</div>",
                unsafe_allow_html=True,
            )

        # ── Step 3: Confirm & Execute ─────────────────────────────────────────
        can_proceed = (
            ("specific student" in delete_mode and len(students_to_delete) > 0)
            or ("entire date" in delete_mode)
        )

        if can_proceed:
            st.markdown("<div class='section-title'>✅ Step 3 — Confirm Deletion</div>", unsafe_allow_html=True)

            # Build a human-readable summary of what will be deleted
            if "specific student" in delete_mode:
                scope_desc = f"**{', '.join(students_to_delete)}** on **{selected_del_date_str}**"
            else:
                scope_desc = f"**all {len(df_sel_day)} records** on **{selected_del_date_str}**"

            st.warning(f"⚠️ You are about to permanently delete {scope_desc} from the attendance file.")

            col_confirm, col_btn = st.columns([2, 1])
            with col_confirm:
                confirmed_del_att = st.checkbox(
                    "I understand this is permanent and cannot be undone",
                    key="del_att_confirm",
                )
            with col_btn:
                do_delete = st.button(
                    "🗑️ Confirm Delete",
                    key="del_att_btn",
                    disabled=not confirmed_del_att,
                    type="primary",
                )

            if do_delete and confirmed_del_att:
                names_arg = students_to_delete if "specific student" in delete_mode else []
                with st.spinner("Deleting records from Excel file…"):
                    ok, msg, count = _delete_attendance_records(selected_del_date, names_arg)

                if ok:
                    st.success(f"✅ {msg}")
                    st.cache_data.clear()   # force fresh reload from Excel
                    st.rerun()
                else:
                    st.error(f"❌ {msg}")

        # ── Live preview of remaining records ─────────────────────────────────
        if enrolled and not df_sel_day.empty:
            st.divider()
            st.markdown(
                "<div class='section-title'>📊 Full Class Status for This Date</div>",
                unsafe_allow_html=True,
            )
            present_names_day = set(df_sel_day["Student Name"].str.lower())
            class_rows = []
            for s in enrolled:
                is_p = s["name"].lower() in present_names_day
                match = df_sel_day[df_sel_day["Student Name"].str.lower() == s["name"].lower()]
                t = match["Time"].values[0] if not match.empty else "—"
                class_rows.append({
                    "Roll No":      s["roll_no"],
                    "Student Name": s["name"],
                    "Status":       "✅ Present" if is_p else "❌ Absent",
                    "Check-in Time": t,
                })
            class_df = pd.DataFrame(class_rows)
            st.dataframe(
                class_df.style.map(
                    lambda v: "color:#00C853;font-weight:600" if "Present" in str(v)
                              else ("color:#FF1744;font-weight:600" if "Absent" in str(v) else ""),
                    subset=["Status"],
                ),
                use_container_width=True,
                hide_index=True,
            )


# ═══════════════════════════════════════════════════════════════════════════════
#  PAGE: MANAGE STUDENTS
# ═══════════════════════════════════════════════════════════════════════════════

elif page == "⚙️ Manage Students":
    st.markdown("<h1 style='color:#fff;margin-bottom:0'>⚙️ Manage Students</h1>", unsafe_allow_html=True)
    st.markdown("<p style='color:#8888aa'>Delete or edit enrolled student records.</p>", unsafe_allow_html=True)

    if not enrolled:
        st.warning("No students enrolled yet. Use the main Tkinter app to enroll students.")
    else:
        # ── Student roster table ──────────────────────────────────────────────
        st.markdown("<div class='section-title'>🎓 Enrolled Students</div>", unsafe_allow_html=True)

        roster_df = pd.DataFrame(enrolled).rename(columns={"name": "Student Name", "roll_no": "Roll No"})
        # Add attendance % column
        if not df_all.empty:
            all_sess = df_all["Date"].dropna().nunique()
            def _pct(name):
                p = df_all[df_all["Student Name"].str.lower() == name.lower()]["Date"].nunique()
                return f"{round(p/all_sess*100,1)}%" if all_sess else "N/A"
            roster_df["Attendance %"] = roster_df["Student Name"].apply(_pct)
        else:
            roster_df["Attendance %"] = "N/A"

        st.dataframe(roster_df.reset_index(drop=True), use_container_width=True, hide_index=True)

        st.divider()

        # ── Two-column layout: Delete | Edit ──────────────────────────────────
        col_del, col_edit = st.columns(2, gap="large")

        # ─── DELETE ──────────────────────────────────────────────────────────
        with col_del:
            st.markdown("<div class='section-title'>🗑️ Delete Student</div>", unsafe_allow_html=True)
            st.markdown(
                "<small style='color:#FF6D00'>⚠️ This permanently removes the student's face data "
                "and retrains the model. Attendance history is kept unless you choose to purge it.</small>",
                unsafe_allow_html=True,
            )
            st.markdown("<br>", unsafe_allow_html=True)

            del_name = st.selectbox(
                "Select student to delete",
                [s["name"] for s in enrolled],
                key="del_select",
            )

            purge_att = st.checkbox(
                "Also delete this student's attendance records",
                value=True,
                key="del_purge",
                help="Removes all historical attendance rows for this student from the Excel file.",
            )

            confirmed_del = st.checkbox(
                f'I confirm I want to remove **{del_name}** from the system',
                key="del_confirm",
            )

            if st.button("🗑️ Delete Student", key="del_btn", disabled=not confirmed_del):
                with st.spinner(f"Deleting {del_name} and retraining model…"):
                    ok, msg = _delete_student_dashboard(del_name, purge_attendance=purge_att)
                if ok:
                    st.success(f"✅ {msg}")
                    st.cache_data.clear()
                    st.rerun()
                else:
                    st.error(f"❌ {msg}")

        # ─── EDIT / RENAME ────────────────────────────────────────────────────
        with col_edit:
            st.markdown("<div class='section-title'>✏️ Edit Student Info</div>", unsafe_allow_html=True)
            st.markdown(
                "<small style='color:#8888aa'>Change the student's name or roll number. "
                "Face model labels and attendance history are updated automatically.</small>",
                unsafe_allow_html=True,
            )
            st.markdown("<br>", unsafe_allow_html=True)

            edit_student = st.selectbox(
                "Select student to edit",
                [s["name"] for s in enrolled],
                key="edit_select",
            )
            edit_obj = next((s for s in enrolled if s["name"] == edit_student), None)

            if edit_obj:
                new_name = st.text_input(
                    "New Name",
                    value=edit_obj["name"],
                    key="edit_name",
                )
                new_roll = st.text_input(
                    "New Roll No",
                    value=edit_obj["roll_no"],
                    key="edit_roll",
                )

                name_changed = new_name.strip() != edit_obj["name"]
                roll_changed = new_roll.strip() != edit_obj["roll_no"]
                has_changes  = (name_changed or roll_changed) and new_name.strip()

                if not new_name.strip():
                    st.warning("Name cannot be empty.")

                if st.button("💾 Save Changes", key="edit_btn", disabled=not has_changes):
                    with st.spinner("Updating student info…"):
                        ok, msg = _rename_student(
                            edit_obj["name"], edit_obj["roll_no"],
                            new_name.strip(), new_roll.strip()
                        )
                    if ok:
                        st.success(f"✅ {msg}")
                        st.cache_data.clear()
                        st.rerun()
                    else:
                        st.error(f"❌ {msg}")
                elif has_changes:
                    changes = []
                    if name_changed: changes.append(f"Name: **{edit_obj['name']}** → **{new_name.strip()}**")
                    if roll_changed: changes.append(f"Roll No: **{edit_obj['roll_no']}** → **{new_roll.strip()}**")
                    st.info("Pending changes:\n" + "\n".join(f"• {c}" for c in changes))
