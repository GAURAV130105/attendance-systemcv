"""
Attendance Logger Module
Handles reading/writing ALL attendance records to a single Excel workbook.

Structure
---------
One file: attendance_records/attendance_records.xlsx
Columns : S.No | Roll No | Student Name | Date | Time | Status

Each row is one attendance event.  A student can appear multiple times across
different dates.  Within a single session (same calendar day) they are only
marked once.
"""
import os
import threading
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from config import ATTENDANCE_DIR, ATTENDANCE_FILE


# ── Style helpers ────────────────────────────────────────────────────────────

def _thin_border():
    s = Side(style='thin')
    return Border(left=s, right=s, top=s, bottom=s)

def _header_style():
    return (
        Font(name='Calibri', bold=True, size=12, color="FFFFFF"),
        PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid"),
        Alignment(horizontal="center", vertical="center"),
    )

def _data_style():
    return (
        Alignment(horizontal="center", vertical="center"),
        PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"),
    )


# ── AttendanceLogger ─────────────────────────────────────────────────────────

class AttendanceLogger:
    """
    Manages the single master attendance workbook.

    ``marked_names`` only tracks marks made **today** so that:
      - No duplicate within the same session.
      - The same student CAN be marked again on a future day.
    """

    def __init__(self):
        self.file_path  = ATTENDANCE_FILE
        # NOTE: date_str is NOT cached at init — we always read the live date
        # so that sessions running past midnight or across days stay correct.
        self._last_date = datetime.now().strftime("%Y-%m-%d")
        # Only track TODAY's marks so future sessions start fresh
        self.marked_names: set = set()

        # ── In-memory cache ───────────────────────────────────────────
        # Holds dicts {name, roll_no, time} for the CURRENT session.
        # get_present_list() now returns this directly (O(1)), rather than
        # scanning the entire Excel sheet on every UI refresh.
        self._present_cache: list[dict] = []

        os.makedirs(ATTENDANCE_DIR, exist_ok=True)

        if os.path.exists(self.file_path):
            self.wb = load_workbook(self.file_path)
            self.ws = self.wb.active
            self._load_todays_marks()
        else:
            self._create_new_workbook()

        # ── Auto-save thread ──────────────────────────────────────────
        # Persists the workbook to disk every 60 seconds so attendance is
        # never lost if the app is closed unexpectedly.
        self._stop_autosave = threading.Event()
        self._autosave_thread = threading.Thread(
            target=self._autosave_loop, daemon=True
        )
        self._autosave_thread.start()

    # ── Workbook setup ───────────────────────────────────────────────────

    def _create_new_workbook(self):
        """Create the master workbook with styled headers."""
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Attendance Records"

        header_font, header_fill, header_align = _header_style()
        border = _thin_border()

        headers    = ["S.No", "Roll No", "Student Name", "Date", "Time", "Status"]
        col_widths = [8,       15,        25,             15,     12,     12]

        for col, (hdr, width) in enumerate(zip(headers, col_widths), 1):
            cell = self.ws.cell(row=1, column=col, value=hdr)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = header_align
            cell.border    = border
            self.ws.column_dimensions[chr(64 + col)].width = width

        self.ws.freeze_panes = "A2"

    @property
    def date_str(self) -> str:
        """Always return today's date string — never stale, even across midnight."""
        return datetime.now().strftime("%Y-%m-%d")

    def _check_date_rollover(self):
        """
        If the calendar date has changed since the last mark, clear the
        marked_names set and present cache so the new day starts fresh.
        Called at the top of mark_present() and is_already_marked().
        """
        today = self.date_str
        if today != self._last_date:
            self._last_date = today
            self.marked_names.clear()
            self._present_cache.clear()
            print(f"[AttendanceLogger] Date rolled over to {today} — session reset.")

    def _load_todays_marks(self):
        """
        Scan the workbook and populate ``marked_names`` with students
        who were already recorded **today** (column D == today's date).
        Also rebuilds _present_cache so get_present_list() is correct after
        a restart mid-session.
        """
        today = self.date_str
        for row in self.ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[2]:
                continue
            row_date = str(row[3]).strip() if row[3] else ""
            if row_date == today:
                name    = str(row[2]).strip()
                roll_no = str(row[1]).strip() if row[1] else ""
                time_v  = str(row[4]).strip() if row[4] else ""
                self.marked_names.add(name)
                self._present_cache.append({
                    "name":    name,
                    "roll_no": roll_no,
                    "time":    time_v,
                })

    # ── Public API ───────────────────────────────────────────────────────

    def mark_present(self, name: str, roll_no: str) -> bool:
        """
        Record *name* as present for today (no duplicates within the same day).

        The mark is stored immediately in the in-memory cache (_present_cache)
        so that get_present_list() returns updated data instantly.  The actual
        Excel write is deferred to close() (or the auto-save thread) so that
        no blocking I/O happens on the recognition hot-path.

        Returns
        -------
        True  – newly marked
        False – already marked today
        """
        # Auto-reset if the calendar date has changed (handles overnight sessions)
        self._check_date_rollover()

        if name in self.marked_names:
            return False

        now = datetime.now()
        self.marked_names.add(name)
        self._present_cache.append({
            "name":    name,
            "roll_no": roll_no,
            "time":    now.strftime("%H:%M:%S"),
            "date":    now.strftime("%Y-%m-%d"),
        })
        return True

    def is_already_marked(self, name: str) -> bool:
        """Return True if this student has already been marked present today."""
        self._check_date_rollover()
        return name in self.marked_names

    def get_present_list(self) -> list:
        """
        Return students marked present in the current session (today).

        O(1) — returns the pre-built in-memory cache directly; no sheet scan.

        Returns
        -------
        list of dicts: {name, roll_no, time}
        """
        return list(self._present_cache)

    # ── Private helpers ───────────────────────────────────────────────────

    def _flush_to_sheet(self):
        """
        Write all in-memory present records to the worksheet.
        Called by save() / close() — never on the hot recognition path.
        Groups entries by their stored date so multi-day sessions are written correctly.
        """
        if not self._present_cache:
            return

        # Collect which dates appear in the cache
        cache_dates = set(e.get("date", self._last_date) for e in self._present_cache)

        # Remove existing rows for those dates to prevent duplicates
        rows_to_delete = [
            row[0].row for row in self.ws.iter_rows(min_row=2)
            if row[3].value and str(row[3].value).strip() in cache_dates
        ]
        for r in sorted(rows_to_delete, reverse=True):
            self.ws.delete_rows(r)

        data_align, present_fill = _data_style()
        border = _thin_border()

        for entry in self._present_cache:
            entry_date = entry.get("date", self._last_date)
            row_num = self.ws.max_row + 1
            sno     = row_num - 1
            row_data = [
                sno,
                entry["roll_no"],
                entry["name"],
                entry_date,
                entry["time"],
                "Present",
            ]
            for col, value in enumerate(row_data, 1):
                cell = self.ws.cell(row=row_num, column=col, value=value)
                cell.alignment = data_align
                cell.border    = border
                cell.fill      = present_fill

    def _autosave_loop(self, interval: int = 60):
        """Background daemon: flush + persist the workbook every `interval` seconds."""
        while not self._stop_autosave.wait(timeout=interval):
            try:
                self._flush_to_sheet()
                self.wb.save(self.file_path)
            except Exception as e:
                print(f"[AttendanceLogger] Auto-save failed: {e}")

    def get_summary(self):
        """
        Return a summary dict for the current session.

        Returns
        -------
        dict: {total_present, date, file_path}
        """
        return {
            "total_present": len(self.marked_names),
            "date":          self.date_str,
            "file_path":     self.file_path,
        }

    def save(self):
        """Flush in-memory records to the sheet and persist to disk."""
        self._flush_to_sheet()
        self.wb.save(self.file_path)

    def close(self):
        """Flush, save, and stop the auto-save thread."""
        self._stop_autosave.set()
        self._flush_to_sheet()
        self.wb.save(self.file_path)
        self.wb.close()


# ── Module-level history helper ──────────────────────────────────────────────

def get_student_history(name: str, roll_no: str = "",
                        extra_present_dates: list = None) -> dict:
    """
    Read the single master attendance workbook and compute per-student stats.

    Scans:
      - Every **unique date** that appears anywhere in the file = a "session day".
      - Dates where the student has a row = present.
      - Remaining session dates = absent.

    Args
    ----
    name                 : Student name (matched against column C, case-insensitive).
    roll_no              : Roll number  (fallback match on column B).
    extra_present_dates  : Additional dates (YYYY-MM-DD strings) to inject as
                           present for this student — used to include in-memory
                           marks that have not yet been flushed to Excel.

    Returns
    -------
    dict with keys:
        total_days, present_days, absent_days, pct,
        dates_present (sorted list), dates_absent (sorted list)
    """
    extra_present_dates = extra_present_dates or []

    all_dates:     set = set()   # every session date in the file
    dates_present: set = set()   # dates where THIS student was present

    if os.path.exists(ATTENDANCE_FILE):
        try:
            wb = load_workbook(ATTENDANCE_FILE, read_only=True, data_only=True)
            ws = wb.active

            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[2]:
                    continue

                row_date = str(row[3]).strip() if row[3] else ""
                row_name = str(row[2]).strip() if row[2] else ""
                row_roll = str(row[1]).strip() if row[1] else ""

                if row_date:
                    all_dates.add(row_date)

                matched = (
                    row_name.lower() == name.lower()
                    or (roll_no and row_roll.lower() == roll_no.lower())
                )
                if matched and row_date:
                    dates_present.add(row_date)

            wb.close()
        except Exception:
            pass   # fall through to use extra_present_dates only

    # Merge in-memory (unflushed) marks so the popup is immediately accurate
    for d in extra_present_dates:
        all_dates.add(d)
        dates_present.add(d)

    if not all_dates:
        return {
            "total_days":    0,
            "present_days":  0,
            "absent_days":   0,
            "pct":           0.0,
            "dates_present": [],
            "dates_absent":  [],
        }

    dates_absent = sorted(all_dates - dates_present)
    dates_present_list = sorted(dates_present)

    total_days   = len(all_dates)
    present_days = len(dates_present_list)
    absent_days  = len(dates_absent)
    pct = (present_days / total_days * 100) if total_days > 0 else 0.0

    return {
        "total_days":    total_days,
        "present_days":  present_days,
        "absent_days":   absent_days,
        "pct":           round(pct, 1),
        "dates_present": dates_present_list,
        "dates_absent":  dates_absent,
    }
